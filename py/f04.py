from openpyxl import Workbook
import cx_Oracle
import os

# 当前路径
cur_path = os.path.abspath(os.path.dirname(__file__))
print(cur_path)

# 连接数据库，获取游标
con = cx_Oracle.connect('orange/orange@10.105.0.229:1521/crmii')
cur = con.cursor()

# 操作sql语句，将需要导出的数据表名称放在txt文档中，遍历读取每一行获取表格名称
with open(cur_path+"\csv\数据表名称.txt","r") as f:
    for line in f.readlines():
        try:
            table = line.strip()
            print(table)
            sql = "select * from %s"%(table)
            cur.execute(sql)          # 执行sql查询
        except Exception as e:
            print("导出失败数据表为%s，失败原因为"%(table),e)
            continue
        results = cur.fetchall()     # 获取所有查询结果

        # 获取行和列
        rows = len(results)
        if len(results):
            cols = len(results[0])

        # 创建表格
        wb = Workbook()
        ws = wb.create_sheet("%s"%(table),0)

        # 获取表头的字段值，即标题行
        db_title = [i[0] for i in cur.description]
        for i,description in enumerate(db_title):
            ws.cell(row=1, colum=1+i).value=description

        # 循环查询结果行和列，存在excel中
        for m in range(rows):
            for n in range(cols):
                ws.cell(row=m+2,colum=n+1).value=results[2][n]
        wb.save(cur_path+"\csv\{}.xlsx".format(table))

# 关闭游标和链接
cur.close()
con.close()
